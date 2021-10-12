"""Process A/V dataset."""

import os
import os.path as op
import time
import warnings

import numpy as np
from scipy import signal, stats
import matplotlib.pyplot as plt
import openpyxl

import pandas as pd
with warnings.catch_warnings(record=True):
    warnings.simplefilter('ignore', FutureWarning)
    from nilearn.glm.first_level import \
        make_first_level_design_matrix, compute_regressor  # noqa
import statsmodels.formula.api as smf
import mne_nirs.preprocessing
import mne_nirs.statistics
import mne_nirs.utils
import mne_nirs.statistics
import mne
from mne.preprocessing.nirs import tddr
subjects = (
    '6003 6005 6006 6007 6008 6009 6010 6011 6012 6013 '
    '6014 6016 6017 6018 6019 6020 6021 6022 6023 6024 '
    '6025 6026 6027 6029').split()
assert len(subjects) == 24

conditions = ('A', 'V', 'AV', 'W')
colors = dict(  # https://personal.sron.nl/~pault/data/colourschemes.pdf
    A='#4477AA',  # blue
    AV='#CCBB44',  # yellow
    V='#EE7733',  # orange
    W='#AA3377',  # purple
)
exp_name = 'av'
runs = tuple(range(1, 3))
duration = 1.8
design = 'event'
plot_subject = '6006'
plot_run = 1
beh_title, beh_idx = 'AV', 0
filt_kwargs = dict(
    l_freq=0.02, l_trans_bandwidth=0.02,
    h_freq=0.2, h_trans_bandwidth=0.02)

run_h = True  # regenerate HbO/HbR

n_jobs = 4  # for GLM

raw_path = 'data'
behavioral_path = op.join('data', 'NIRx behavioral data.xlsx')
proc_path = 'processed'
results_path = 'results'
os.makedirs(results_path, exist_ok=True)
os.makedirs(proc_path, exist_ok=True)
use = None
all_sci = list()
for subject in subjects[0 if run_h else subjects.index(plot_subject):]:
    for run in runs:
        root = f'AV{run}' if run < 3 else 'McGurk'
        fname = op.join(raw_path, root, subject)
        base = f'{subject}_{run:03d}'
        base_pr = base.ljust(20)
        if not run_h:
            if subject != plot_subject or run != plot_run:
                continue
        raw_intensity = mne.io.read_raw_nirx(fname)
        raw_od = mne.preprocessing.nirs.optical_density(
            raw_intensity, verbose='error')
        # good/bad channels
        peaks = np.ptp(raw_od.get_data('fnirs'), axis=-1)
        flat_names = [
            raw_od.ch_names[f].split(' ')[0]
            for f in np.where(peaks < 0.001)[0]]
        sci = mne.preprocessing.nirs.scalp_coupling_index(raw_od)
        all_sci.extend(sci)
        sci_mask = (sci < 0.25)
        got = np.where(sci_mask)[0]
        print(f'    Run {base_pr}: {len(got)}/{len(raw_od.ch_names)} bad')
        # assign bads
        assert raw_od.info['bads'] == []
        bads = set(raw_od.ch_names[pick] for pick in got)
        bads = bads | set(ch_name for ch_name in raw_od.ch_names
                          if ch_name.split(' ')[0] in flat_names)
        bads = sorted(bads)
        raw_tddr = tddr(raw_od)
        raw_tddr_bp = raw_tddr.copy().filter(**filt_kwargs)
        raw_tddr_bp.info['bads'] = bads
        picks = mne.pick_types(raw_tddr_bp.info, fnirs=True)
        peaks = np.ptp(raw_tddr_bp.get_data(picks), axis=-1)
        assert (peaks > 1e-5).all()
        raw_tddr_bp.info['bads'] = []
        raw_h = mne.preprocessing.nirs.beer_lambert_law(raw_tddr_bp, 6.)
        # wait until now to assign bads so that we can choose later whether
        # we want the MATLAB bads or the Python ones
        h_bads = [
            ch_name for ch_name in raw_h.ch_names
            if ch_name.split(' ')[0] in set(bad.split(' ')[0] for bad in bads)]
        assert len(bads) == len(h_bads)
        raw_h.info['bads'] = h_bads
        raw_h.info._check_consistency()
        picks = mne.pick_types(raw_h.info, fnirs=True)
        peaks = np.ptp(raw_h.get_data(picks), axis=-1)
        assert (peaks > 1e-9).all()  # TODO: Maybe too small
        raw_h.save(op.join(proc_path, f'{base}_hbo_raw.fif'),
                   overwrite=True)
        if subject == plot_subject and run == plot_run:
            assert use is None
            use = dict(intensity=raw_intensity,
                       od=raw_od,
                       tddr=raw_tddr,
                       h=raw_h,
                       run=run)
        del raw_intensity, raw_od, raw_tddr, raw_tddr_bp, raw_h
assert isinstance(use, dict)
ch_names = [ch_name.rstrip(' hbo') for ch_name in use['h'].ch_names[::2]]
info = use['h'].info

###############################################################################
# Settings

plt.rcParams['axes.titlesize'] = 8
plt.rcParams['axes.labelsize'] = 8
plt.rcParams['xtick.labelsize'] = 8
plt.rcParams['ytick.labelsize'] = 8

###############################################################################
# Channel example figure

sfreq = 7.8125  # all analysis at this rate


def _make_design(raw_h, design, subject=None, run=None):
    events, _ = mne.events_from_annotations(raw_h)
    # mis-codings
    if subject == '6011' and run == 1:
        events[1, 2] = 5
    elif subject == '6014' and run == 2:
        events[97, 2] = 5
    n_times = len(raw_h.times)
    stim = np.zeros((n_times, 4))
    events = events[events[:, 2] != 1]
    events[:, 2] -= 1
    assert len(events) == 100, len(events)
    if subject == '6010' and run == 2:
        print('fixing events ...', end='')
        events[42, 2] = 2
    want = [0] + [25] * 4
    count = np.bincount(events[:, 2])
    assert np.array_equal(count, want), count
    assert events.shape == (100, 3), events.shape
    # mne.viz.plot_events(events)
    if np.diff(events[:2, 0]) > 50:
        print('fixing timing ...', end='')
        assert (subject, run) in (('6003', 1), ('6006', 1), (None, None)), (subject, run)  # noqa: E501
        events[0::5, 0] = events[1::5, 0] - 31
    # mne.viz.plot_events(events)
    if design == 'block':
        events = events[0::5]
        duration = 20.
        assert np.array_equal(np.bincount(events[:, 2]), [0] + [5] * 4)
    else:
        assert design == 'event'
        assert len(events) == 100
        duration = 1.8
        assert events.shape == (100, 3)
        events_r = events[:, 2].reshape(20, 5)
        assert (events_r == events_r[:, :1]).all()
        del events_r
    idx = (events[:, [0, 2]] - [0, 1]).T
    assert np.in1d(idx[1], np.arange(len(conditions))).all()
    stim[tuple(idx)] = 1
    assert raw_h.info['sfreq'] == sfreq  # necessary for below logic to work
    n_block = int(np.ceil(duration * sfreq))
    stim = signal.fftconvolve(stim, np.ones((n_block, 1)), axes=0)[:n_times]
    dm_events = pd.DataFrame({
        'trial_type': [conditions[ii] for ii in idx[1]],
        'onset': idx[0] / raw_h.info['sfreq'],
        'duration': n_block / raw_h.info['sfreq']})
    dm = make_first_level_design_matrix(
        raw_h.times, dm_events, hrf_model='glover',
        drift_model='polynomial', drift_order=0)
    return stim, dm, events


###############################################################################
# Plot the design matrix and some raw traces

fig, axes = plt.subplots(2, 1, figsize=(6., 3), constrained_layout=True)
# Design
ax = axes[0]
raw_h = use['h']
stim, dm, _ = _make_design(raw_h, design)
for ci, condition in enumerate(conditions):
    color = colors[condition]
    ax.fill_between(
        raw_h.times, stim[:, ci], 0, edgecolor='none', facecolor='k',
        alpha=0.5)
    model = dm[conditions[ci]].to_numpy()
    ax.plot(raw_h.times, model, ls='-', lw=1, color=color)
    x = raw_h.times[np.where(model > 0)[0][0]]
    ax.text(
        x + 10, 1.1, condition, color=color, fontweight='bold', ha='center')
ax.set(ylabel='Modeled\noxyHb', xlabel='', xlim=raw_h.times[[0, -1]])

# HbO/HbR
ax = axes[1]
picks = [pi for pi, ch_name in enumerate(raw_h.ch_names)
         if 'S4_D4' in ch_name]
assert len(picks) == 2
colors = dict(hbo='r', hbr='b')
ylim = np.array([-0.5, 0.5])
for pi, pick in enumerate(picks):
    color = colors[raw_h.ch_names[pick][-3:]]
    data = raw_h.get_data(pick)[0] * 1e6
    val = np.ptp(data)
    assert val > 0.01
    ax.plot(raw_h.times, data, color=color, lw=1.)
ax.set(ylim=ylim, xlabel='Time (s)', ylabel='μM',
       xlim=raw_h.times[[0, -1]])
del raw_h
for ax in axes:
    for key in ('top', 'right'):
        ax.spines[key].set_visible(False)
for ext in ('png', 'svg'):
    fig.savefig(
        op.join(
            results_path, f'figure_1_{exp_name}.{ext}'))


###############################################################################
# Run GLM analysis and epoching

df_cha = pd.DataFrame()
for subject in subjects:
    fname = op.join(proc_path, f'{subject}_{exp_name}.h5')
    if not op.isfile(fname):
        subj_cha = pd.DataFrame()
        t0 = time.time()
        print(f'Running GLM for {subject}... ', end='')
        for run in runs:
            base = f'{subject}_{run:03d}'
            raw_h = mne.io.read_raw_fif(
                op.join(proc_path, f'{base}_hbo_raw.fif'))
            if raw_h.info['sfreq'] == sfreq / 2.:
                print('resampling... ', end='')
                raw_h.resample(sfreq)
            assert raw_h.info['sfreq'] == sfreq, raw_h.info['sfreq']
            _, dm, _ = _make_design(raw_h, design, subject, run)
            glm_est = mne_nirs.statistics.run_glm(
                raw_h, dm, noise_model='ols', n_jobs=n_jobs)
            cha = glm_est.to_dataframe()
            cha['subject'] = subject
            cha['run'] = run
            # add good/badness of the channel
            cha['good'] = ~np.in1d(cha['ch_name'], bads)
            subj_cha = subj_cha.append(cha)
            del raw_h
        subj_cha.to_hdf(fname, 'subj_cha', mode='w')
        print(f'{time.time() - t0:0.1f} sec')
    df_cha = df_cha.append(pd.read_hdf(fname))
df_cha.reset_index(drop=True, inplace=True)

# block averages
event_id = {condition: ci for ci, condition in enumerate(conditions, 1)}
evokeds = {condition: dict() for condition in conditions}
for subject in subjects:
    fname = op.join(
        proc_path, f'{subject}-{exp_name}-ave.fif')
    if not op.isfile(fname):
        tmin, tmax = -2, 38
        baseline = (None, 0)
        t0 = time.time()
        print(f'Creating block average for {subject} ... ', end='')
        raws = list()
        events = list()
        for run in runs:
            base = f'{subject}_{run:03d}'
            raw_h = mne.io.read_raw_fif(
                op.join(proc_path, f'{base}_hbo_raw.fif'))
            if raw_h.info['sfreq'] == sfreq / 2:
                raw_h.resample(sfreq)
            assert raw_h.info['sfreq'] == sfreq
            events.append(_make_design(raw_h, 'block', subject, run)[2])
            raws.append(raw_h)
        bads = sorted(set(sum((r.info['bads'] for r in raws), [])))
        for r in raws:
            r.info['bads'] = bads
        raw_h, events = mne.concatenate_raws(raws, events_list=events)
        epochs = mne.Epochs(raw_h, events, event_id, tmin=tmin, tmax=tmax,
                            baseline=baseline)
        this_ev = [epochs[condition].average() for condition in conditions]
        assert all(ev.nave > 0 for ev in this_ev)
        mne.write_evokeds(fname, this_ev)
        print(f'{time.time() - t0:0.1f} sec')
    for condition in conditions:
        evokeds[condition][subject] = mne.read_evokeds(fname, condition)


# Get behavioral data
beh = openpyxl.load_workbook(behavioral_path).worksheets[0]
assert beh.cell(1, 7).value == 'pMcGurk'
beh_kinds = ('="-9dB SNR ii"', '="-6dB SNR ii"', 'pMcGurk')
beh_short = {
    '="-6dB SNR ii"': '-6',
    '="-9dB SNR ii"': '-9',
    'pMcGurk': 'pM',
}
for bi, b in enumerate(beh_kinds, 5):
    assert beh.cell(1, bi).value == b, b
behs = dict()
for ri in range(2, 10000):
    subject = beh.cell(ri, 1).value
    if subject is None:
        break
    subject = str(int(subject))
    if subject == '6030':
        continue  # not used
    if subject == '6006' and exp_name == 'mcgurk':
        continue  # excluded
    assert subject in subjects
    behs[subject] = dict((b, beh.cell(ri, bi).value)
                         for bi, b in enumerate(beh_kinds, 5))
    if subject == '6023':
        behs['6023']['="-6dB SNR ii"'] = np.nan
    behs[subject] = dict((key, float(val)) for key, val in behs[subject].items())
assert set(behs) == set(subjects)

# Exclude bad channels
bad = dict()
for subject in subjects:
    for run in runs:
        base = f'{subject}_{run:03d}'
        this_info = mne.io.read_info(
            op.join(proc_path, f'{base}_hbo_raw.fif'))
        bad[(subject, run)] = sorted(
            this_info['ch_names'].index(bad) for bad in this_info['bads'])
    assert np.in1d(bad[(subject, run)], np.arange(len(use['h'].ch_names))).all()  # noqa: E501
# make life easier by combining across runs
bad_combo = dict()
for (subject, run), bb in bad.items():
    bad_combo[subject] = sorted(set(bad_combo.get(subject, [])) | set(bb))
bad = bad_combo
assert set(bad) == set(subjects)

start = len(df_cha)
n_drop = 0
for subject, bb in bad.items():
    if not len(bb):
        continue
    drop_names = [use['h'].ch_names[b] for b in bb]
    is_subject = (df_cha['subject'] == subject)
    assert len(is_subject) == len(df_cha)
    # is_run = (df_cha['run'] == run)
    drop = df_cha.index[
        is_subject &
        # is_run &
        np.in1d(df_cha['ch_name'], drop_names)]
    n_drop += len(drop)
    if len(drop):
        print(f'Dropping {len(drop)} for {subject}')  # {run}')
        df_cha.drop(drop, inplace=True)
end = len(df_cha)
assert n_drop == start - end, (n_drop, start - end)
# combine runs by averaging estimates
sorts = ['subject', 'ch_name', 'Chroma', 'Condition', 'run']
df_cha.sort_values(
    sorts, inplace=True)
assert (np.array(df_cha['run']).reshape(-1, 2) == runs).all()
theta = np.array(df_cha['theta']).reshape(-1, len(runs)).mean(-1)
df_cha.drop(
    [col for col in df_cha.columns if col not in sorts[:-1]], axis='columns',
    inplace=True)
df_cha.reset_index(drop=True, inplace=True)
df_cha = df_cha[::len(runs)]
df_cha.reset_index(drop=True, inplace=True)
df_cha['theta'] = theta


def _mixed_df(ch_summary):
    ch_model = smf.mixedlm(  # remove intercept, interaction between ch+cond
        "theta ~ -1 + ch_name:Condition",
        ch_summary, groups=ch_summary["subject"]).fit(method='powell')
    ch_model_df = mne_nirs.statistics.statsmodels_to_results(ch_model)
    ch_model_df['P>|z|'] = ch_model.pvalues
    ch_model_df.drop([idx for idx in ch_model_df.index if '[constant]' in idx],
                     inplace=True)
    return ch_model_df


times = evokeds[conditions[0]][subjects[0]].times
info = evokeds[conditions[0]][subjects[0]].info

# Run group level model and convert to dataframe
use_lim = [0, 100]  # [0, 100]
lims = np.percentile([b['pMcGurk'] for b in behs.values()], use_lim)
use_subjects = [subj for subj in subjects
                if lims[0] <= behs[subj]['pMcGurk'] <= lims[1]]
ch_summary = df_cha.query("Chroma in ['hbo']").copy()
ch_summary_use = df_cha.query(
    f"Chroma in ['hbo'] and subject in {use_subjects}").copy()
ch_model_df = _mixed_df(ch_summary_use)
print(f'Correcting for {len(ch_model_df["P>|z|"])} comparisons using FDR')
assert len(ch_model_df['P>|z|']) == len(ch_names) * len(conditions)

_, ch_model_df['P_fdr'] = mne.stats.fdr_correction(
    ch_model_df['P>|z|'], method='indep')
sig_chs = dict()
zs = dict()
for condition in conditions:
    sig_df = ch_model_df[
        (ch_model_df['P_fdr'] < 0.05) &
        (ch_model_df['Condition'] == condition)]
    sig_chs[condition] = sorted(
        (use['h'].ch_names.index(row[1]['ch_name']), row[1]['P_fdr'])
        for row in sig_df.iterrows())
    ch_model_df[ch_model_df['Condition'] == condition]
    zs[condition] = np.array([
        ch_model_df[(ch_model_df['Condition'] == condition) &
                    (ch_model_df['ch_name'] == ch_name)]['z'][0]
        for ch_name in info['ch_names'][::2]], float)
    assert zs[condition].shape == (52,)
    assert np.isfinite(zs[condition]).all()


def _plot_sig_chs(sigs, ax):
    if sigs and isinstance(sigs[0], tuple):
        sigs = [s[0] for s in sigs]
    ch_groups = [sigs, np.setdiff1d(np.arange(info['nchan']), sigs)]
    mne.viz.plot_sensors(
        info, 'topomap', 'hbo', title='', axes=ax,
        show_names=True, ch_groups=ch_groups)
    ax.collections[0].set(lw=0)
    c = ax.collections[0].get_facecolor()
    c[(c[:, :3] == (0.5, 0, 0)).all(-1)] = (0., 0., 0., 0.1)
    c[(c[:, :3] == (0, 0, 0.5)).all(-1)] = (0., 1., 0., 0.5)
    ax.collections[0].set_facecolor(c)
    ch_names = [info['ch_names'][idx] for idx in sigs]
    texts = list(ax.texts)
    got = []
    for text in list(texts):
        try:
            idx = ch_names.index(text.get_text())
        except ValueError:
            text.remove()
        else:
            got.append(idx)
            text.set_text(f'{sigs[idx] // 2 + 1}')
            text.set(fontsize='xx-small', zorder=5, ha='center')
    assert len(got) == len(sigs), (got, list(sigs))


def _plot_sigs(sig_chs, all_corrs=()):
    n_col = max(len(x) for x in sig_chs.values()) + 1
    n_row = len(conditions)
    figsize = (n_col * 1.0, n_row * 1.0)
    fig, axes = plt.subplots(
        n_row, n_col, figsize=figsize, constrained_layout=True, squeeze=False)
    h_colors = {0: 'r', 1: 'b'}
    xticks = [0, 10, 20, 30]
    ylim = [-0.1, 0.15]
    yticks = [-0.1, -0.05, 0, 0.05, 0.1]
    xlim = [times[0], 35]
    ylim = np.array(ylim)
    yticks = np.array(yticks)
    for ci, condition in enumerate(conditions):
        ii = 0
        sigs = sig_chs[condition]
        if len(sigs) == 0:
            sigs = [(None, None)]
        for ii, (ch_idx, ch_p) in enumerate(sigs):
            ax = axes[ci, ii]
            if ch_idx is not None:
                for jj in range(2):  # HbO, HbR
                    color = h_colors[jj]
                    a = 1e6 * np.array(
                        [evokeds[condition][subject].data[ch_idx + jj]
                         for subject in use_subjects
                         if ch_idx + jj not in bad.get(subject, [])], float)
                    m = np.mean(a, axis=0)
                    lower, upper = stats.t.interval(
                        0.95, len(a) - 1, loc=m, scale=stats.sem(a, axis=0))
                    ax.fill_between(
                        times, lower, upper, facecolor=color,
                        edgecolor='none', lw=0, alpha=0.25, zorder=3,
                        clip_on=False)
                    ax.plot(times, m, color=color, lw=1, zorder=4,
                            clip_on=False)
                # Correlations
                this_df = ch_summary_use.query(
                    f'ch_name == {repr(use["h"].ch_names[ch_idx])} and '
                    f'Chroma == "hbo" and '
                    f'Condition == {repr(condition)}')
                assert 8 <= len(this_df) <= len(subjects), len(this_df)
                a = np.array(this_df['theta'])
                cs = list()
                for kind in beh_kinds:
                    b = np.array([behs[subject][kind]
                                  for subject in this_df['subject']])
                    mask = np.isfinite(b)
                    assert 8 <= mask.sum() <= len(subjects)
                    r, p = stats.kendalltau(a[mask], b[mask])
                    if p < 0.05 or kind in all_corrs:
                        cs.append(f'{beh_short[kind]}: τ{r:+0.2f} p{p:0.2f}')
                if len(cs):
                    cs = [''] + cs
                c = '\n'.join(cs)
                ax.text(times[-1], ylim[1],
                        f'ch{ch_idx // 2 + 1}\np={ch_p:0.5f}{c}',
                        ha='right', va='top', fontsize='x-small')
            ax.axvline(20, ls=':', color='0.5', zorder=2, lw=1)
            ax.axhline(0, ls='-', color='k', zorder=2, lw=0.5)
            ax.set(xticks=xticks, yticks=yticks)
            ax.set(xlim=xlim, ylim=ylim)
            for key in ('top', 'right'):
                ax.spines[key].set_visible(False)
            if ax.get_subplotspec().is_last_row():
                ax.set(xlabel='Time (sec)')
            else:
                ax.set_xticklabels([''] * len(xticks))
            if ax.get_subplotspec().is_first_col():
                ax.set_ylabel(condition)
            else:
                ax.set_yticklabels([''] * len(yticks))
            for key in ('top', 'right'):
                ax.spines[key].set_visible(False)
        for ii in range(ii + 1, n_col - 1):
            fig.delaxes(axes[ci, ii])
        # montage
        ax = axes[ci, -1]
        if sigs[0][0] is None:
            fig.delaxes(ax)
        else:
            # plot montage
            _plot_sig_chs(sigs, ax)
    return fig

fig = _plot_sigs(sig_chs)
for ext in ('png', 'svg'):
    fig.savefig(op.join(results_path, f'stats_{exp_name}.{ext}'))

###############################################################################
# Source space projection
info = use['h'].copy().pick_types(fnirs='hbo', exclude=()).info
info['bads'] = []
assert tuple(zs) == conditions
evoked = mne.EvokedArray(np.array(list(zs.values())).T, info)
picks = np.arange(len(evoked.ch_names))
for ch in evoked.info['chs']:
    assert ch['coord_frame'] == mne.io.constants.FIFF.FIFFV_COORD_HEAD
stc = mne.stc_near_sensors(
    evoked, trans='fsaverage', subject='fsaverage', mode='weighted',
    distance=0.02, project=True, picks=picks)
# Split channel indices by left lat, posterior, right lat:
# num_map = {name: str(ii) for ii, name in enumerate(evoked.ch_names)}
# evoked.copy().rename_channels(num_map).plot_sensors(show_names=True)
view_map = [np.arange(19), np.arange(19, 33), np.arange(33, 52)]
surf = mne.read_bem_surfaces(
    mne.utils.get_subjects_dir() +
    '/fsaverage/bem/fsaverage-5120-5120-5120-bem.fif', s_id=1)  # brain
for ci, condition in enumerate(conditions):
    this_sig = [v[0] // 2 for v in sig_chs[condition]]
    assert np.in1d(this_sig, np.arange(52)).all()
    pos = np.array([info['chs'][idx]['loc'][:3] for idx in this_sig])
    pos.shape = (-1, 3)  # can be empty
    # head->MRI
    trans = mne.transforms._get_trans('fsaverage', 'head', 'mri')[0]
    # project to brain surface
    pos = mne.transforms.apply_trans(trans, pos)  # now in MRI coords
    pos = mne.surface._project_onto_surface(pos, surf, project_rrs=True)[2]
    # plot
    brain = stc.plot(hemi='both', views=['lat', 'frontal', 'lat'],
                     initial_time=evoked.times[ci], cortex='low_contrast',
                     time_viewer=False, show_traces=False,
                     surface='pial', smoothing_steps=0, size=(1200, 400),
                     clim=dict(kind='value', pos_lims=[0., 1.25, 2.5]),
                     colormap='RdBu_r', view_layout='horizontal',
                     colorbar=(0, 1), time_label='', background='w',
                     brain_kwargs=dict(units='m'),
                     add_data_kwargs=dict(colorbar_kwargs=dict(
                         title_font_size=24, label_font_size=24, n_labels=5,
                         title='z score')))
    brain.show_view('lat', hemi='lh', row=0, col=0)
    brain.show_view(azimuth=270, elevation=90, row=0, col=1)
    # significant channel white text overlay
    pl = brain.plotter
    used = np.zeros(len(this_sig))
    for vi in range(3):
        this_idx = np.where(np.in1d(this_sig, view_map[vi]))[0]
        assert not used[this_idx].any()
        used[this_idx] = True
        pl.subplot(0, vi)
        vp = pl.renderer  # subclass of vtkViewport
        for idx in this_idx:
            ch_pos = pos[idx]
            vp.SetWorldPoint(np.r_[ch_pos, 1.])
            vp.WorldToDisplay()
            ch_pos = (np.array(vp.GetDisplayPoint()[:2]) -
                      np.array(vp.GetOrigin()))
            actor = pl.add_text(
                str(this_sig[idx] + 1), ch_pos,
                font_size=12, color=(1., 1., 1.))
            prop = actor.GetTextProperty()
            prop.SetVerticalJustificationToCentered()
            prop.SetJustificationToCentered()
            actor.SetTextProperty(prop)
            prop.SetBold(True)
    assert used.all()
    brain.show_view('lat', hemi='rh', row=0, col=2)
    plt.imsave(
        op.join(results_path, f'brain_{exp_name}_{condition}.png'), pl.image)
    brain.close()
